from openpyxl import load_workbook
from dotenv import load_dotenv
from langchain_openai import ChatOpenAI, OpenAIEmbeddings
from langchain_core.prompts import ChatPromptTemplate, MessagesPlaceholder
from langchain.agents import create_openai_functions_agent, AgentExecutor
from langchain_community.tools.tavily_search import TavilySearchResults


def get_material_data(material_list_file) -> dict:
    """
    Reads material data from an Excel file and organizes it into a nested dictionary.

    Args:
        material_list_file (str): Path to the Excel file containing material data.

    Returns:
        dict: A dictionary organized as {category: {subcategory: [grades]}}.
    """
    try:
        mat_wb = load_workbook(material_list_file, read_only=True)
    except FileNotFoundError:
        print(f"Error: File '{material_list_file}' not found.")
        return {}
    except Exception as e:
        print(f"Error: Failed to load workbook: {e}")
        return {}

    mat_ws = mat_wb.active
    material_data_dict: dict = {}

    for row in mat_ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue  # Skip rows with empty grade cells, but don't break the loop

        category = row[5].value
        subcategory = row[6].value
        grade = row[0].value

        if not category or not subcategory or not grade:
            continue  # Skip rows with missing required values

        if category not in material_data_dict:
            material_data_dict[category] = {}
        if subcategory not in material_data_dict[category]:
            material_data_dict[category][subcategory] = []
        if grade not in material_data_dict[category][subcategory]:
            material_data_dict[category][subcategory].append(grade)

    return material_data_dict


def get_categories(material_data: dict) -> list:
    """
    Get the categories from the material data.

    Args:
        material_data (dict): A dictionary containing material data.

    Returns:
        list: A list of categories extracted from the material data.
    """
    return material_data.keys()


def get_subcategories(material_data: dict, category: str) -> list:
    """
    Get the subcategories for a given category from the material data.

    Args:
        material_data (dict): A dictionary containing material data.
        category (str): The category for which to retrieve subcategories.

    Returns:
        list: A list of subcategories for the given category.
    """
    return material_data[category].keys()


def get_grades(material_data: dict, category: str, subcategory: str) -> list:
    """
    Get the grades for a given category and subcategory from the material data.

    Args:
        material_data (dict): A dictionary containing material data.
        category (str): The category of the material.
        subcategory (str): The subcategory of the material.

    Returns:
        list: A list of grades for the given category and subcategory.
    """
    return material_data[category][subcategory]


def get_material_information(material_name: str) -> str:
    """
    Retrieves the description of a material based on its name.

    Args:
        material_name (str): The name of the material.

    Returns:
        str: A description of the material.

    Raises:
        None

    Example:
        >>> get_material_information("steel")
        "Steel is a strong and durable material commonly used in construction and manufacturing. It is known for its high tensile strength and resistance to corrosion."
    """

    # Load API keys
    load_dotenv()

    # Agent setup
    model = ChatOpenAI(model="gpt-3.5-turbo-1106", temperature=0.3)
    prompt = ChatPromptTemplate.from_messages(
        [
            (
                "system",
                "You are a material expert. Please describe the following material, focusing on its type and listing all material equivalents based on different standards. Return a string with a maximum of three sentences.",
            ),
            ("human", "{material}"),
            MessagesPlaceholder(variable_name="agent_scratchpad"),
        ]
    )

    # Search tool
    search = TavilySearchResults(max_results=3)
    tools = [search]

    # Agent
    agent = create_openai_functions_agent(model, tools, prompt)
    agent_executor = AgentExecutor(agent=agent, tools=tools)

    response = agent_executor.invoke({"material": material_name})
    return response


def classify_material(
    material_name: str, material_information: str, options: list
) -> str:
    """
    Classifies the given material based on its name and information.

    Args:
        material_name (str): The name of the material.
        material_information (str): The information about the material.
        options (list): A list of available options to choose from.

    Returns:
        str: The exact name of the chosen option from the list.

    """
    # Load API keys
    load_dotenv()

    # Agent setup
    model = ChatOpenAI(model="gpt-3.5-turbo-1106", temperature=0.1)
    prompt = ChatPromptTemplate.from_messages(
        [
            (
                "system",
                "For the material described below, choose an appropriate option from these options: {options}. Return only the exact name of the option from the list, without any extra words or characters.",
            ),
            ("human", "{material} is described: {material_information}"),
        ]
    )

    # agent = create_openai_functions_agent(llm=model, prompt=prompt, )
    # agent_executor = AgentExecutor(agent=agent)
    chain = prompt | model
    response = chain.invoke(
        {
            "options": options,
            "material": material_name,
            "material_information": material_information,
        }
    )
    return response.content


def correct_answer_from_list(
    answer: str, options: list, counter: int = 0, max_attempts: int = 3
) -> str:
    """Corrects answer so it is chosed from the list

    Args:
        answer (str): with wrong answer
        options (list): list of correct answers
        counter (int): number of attempts
        max_attempts (int): max number of

    Returns:
        str: with corrected answer
        None: when fails
    """

    # Load API keys
    load_dotenv()

    # Checks if the repeated correction counter is not capped
    if counter >= max_attempts:
        print(
            f"correct_answer_from_list could not correct {answer} after {counter} attempts"
        )
        return None
    model = ChatOpenAI(
        model="gpt-3.5-turbo", temperature=0.2, max_tokens=100, verbose=True
    )

    # Prompt template
    prompt = ChatPromptTemplate.from_template(
        "Correct this answer: {answer} so it is one of these options in this list {options}. Return only the exact name of the option from the list, without any extra words or characters.",
    )
    # Create llm chain
    chain = prompt | model
    response = chain.invoke({"answer": answer, "options": options})

    # Error handling when checking output type
    try:
        new_answer = str(response.content)
        if new_answer in options:
            return new_answer
        else:
            counter += 1
            correct_answer_from_list(new_answer, options, counter)
    except TypeError:
        print("correct_answer_from_list got wrong type of answer from llm")
        return None


def start_workflow(material_name: str, material_db: dict):
    """
    Starts the workflow for classifying a material based on its name and information.

    Args:
        material_name (str): The name of the material to be classified.
        material_db (dict): The database containing material information.

    Returns:
        dict: A dictionary containing the classified category, subcategory, and grade of the material.
    """

    # Gather information
    material_info = get_material_information(material_name)

    # Get material category
    categories = get_categories(material_db)
    mat_category = classify_material(
        material_name=material_name,
        material_information=material_info,
        categories=categories,
    )
    # Check if the output is chosen from given options
    if mat_category not in categories:
        mat_category = correct_answer_from_list(mat_category, categories)
    # If correction failed
    if not mat_category:
        mat_category = "Other"

    # Get subcategory
    subcategories = get_subcategories(material_data=material_db, category=mat_category)
    mat_subcategory = classify_material(material_name, material_info, subcategories)
    # Check if the output is chosen from given options
    if mat_subcategory not in subcategories:
        mat_subcategory = correct_answer_from_list(mat_subcategory, subcategories)
    # If correction failed
    if not mat_subcategory:
        mat_subcategory = "Other"

    # Get grade
    grades = get_grades(
        material_data=material_db, category=mat_category, subcategory=mat_subcategory
    )
    mat_grade = classify_material(material_name, material_info, grades)
    # Check if the output is chosen from given options
    if mat_grade not in grades:
        mat_grade = correct_answer_from_list(mat_grade, grades)
    # If correction failed
    if not mat_grade:
        mat_grade = "Other"

    return {
        "category": mat_category,
        "subcategory": mat_subcategory,
        "grade": mat_grade,
    }
